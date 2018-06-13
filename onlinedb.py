import MySQLdb
import django , os, os.path
import click, requests
from bs4 import BeautifulSoup
import urllib.request
os.environ.setdefault('DJANGO_SETTINGS_MODULE', "onlineproject.settings")
django.setup()
from openpyxl import load_workbook
from onlineapp.models import *


@click.group()
@click.pass_context
def cli(ctx):
    pass


@cli.command()
def cleardata():
    Mocktest1.objects.all().delete()
    Student.objects.all().delete()
    College.objects.all().delete()
    print("deleted")


HOSTNAME = "localhost"
USERNAME = "avinash"
USERPWD = "avinash"


def get_mysqldb_connection(HOSTNAME, USERNAME, USERPWD, DATABASE):
    if DATABASE != "":
        return MySQLdb.connect(HOSTNAME, USERNAME, USERPWD, DATABASE)
    else:
        return MySQLdb.connect(HOSTNAME, USERNAME, USERPWD)


@cli.command()
@click.argument("database_name",default=None)
def dropdb(database_name):
    conn = get_mysqldb_connection(HOSTNAME, USERNAME, USERPWD, "")
    cursor = conn.cursor()
    try:
        cursor.execute("drop database "+database_name)
        os.remove("onlineapp/migrations/0001_initial.py")
        print("%s dropped successfully" %database_name)
    except Exception as e:
        print(e)
    conn.close()


@cli.command()
def recreate():
    os.system("python onlinedb.py dropdb mrnd")
    os.system("python onlinedb.py createdb mrnd")
    os.system("python onlinedb.py populatedb students.xlsx test.html")


@cli.command()
@click.argument("database_name",default=None)
def createdb(database_name):
    conn = get_mysqldb_connection(HOSTNAME, USERNAME, USERPWD, "")
    cursor = conn.cursor()
    try:
        cursor.execute("create database "+database_name)

        os.system("python manage.py makemigrations")
        os.system("python manage.py migrate")

        print("%s created successfully" %database_name)

    except Exception as e:
        print(e)




def get_html_data(url_or_file):
    html = ""
    if os.path.isfile(url_or_file):
        if not url_or_file.endswith(".html"):
            raise TypeError("Only HTML files can be processed")
        with open(url_or_file) as f:
            html = f.read()
    else:
        f = urllib.request.urlopen(
            url_or_file)
        html = f.read()
    return html



def isExcel(filename):
    return filename.endswith(".xlsx") or filename.endswith(".xls")


def isHTML(filename):
    return filename.endswith(".html")



def worksheet_to_database_table(college_worksheet, studentsheet,dropped_students, mocktest1_html, worksheet_startrow=1, ):
    max_row = college_worksheet.max_row
    max_col = college_worksheet.max_column
    college_dict = dict()

    # Storing the Colleges details into the database.

    for each_row in range(worksheet_startrow, max_row + 1):
        col_values = []
        for each_col in range(1, max_col + 1):
            col_values.append(college_worksheet.cell(each_row, each_col).value)
        c = College(name=col_values[0], location=col_values[2], acronym=col_values[1], contact = col_values[3])
        c.save()
        college_dict[col_values[1]] = c

    student_dict = dict()

    # Storing the Students excel data into Student table of database.

    for each_row in range(worksheet_startrow,studentsheet.max_row + 1):
        col_values1 = []
        for each_col in range(1,studentsheet.max_column + 1):
            col_values1.append(studentsheet.cell(each_row, each_col).value)
        s = Student(name=col_values1[0],email=col_values1[2],db_folder = col_values1[3].lower(),dropped_out=False,college = college_dict[col_values1[1]])
        s.save()
        student_dict[col_values1[3].lower()] = s

    for each_row in range(worksheet_startrow,dropped_students.max_row + 1):
        col_values1 = []
        for each_col in range(1,dropped_students.max_column + 1):
            col_values1.append(dropped_students.cell(each_row, each_col).value)
        s = Student(name=col_values1[0],email=col_values1[2],db_folder = col_values1[3].lower(),dropped_out=True, college = college_dict[col_values1[1]])
        s.save()
        student_dict[col_values1[3].lower()] = s

    #Converting html code and uploading it to database

    html = get_html_data(mocktest1_html)
    soup = BeautifulSoup(html, "lxml")
    table = soup.find("table")
    rows = table.findAll("tr")
    x, y = 1, 1

    for row in rows:
        if x != 1:
            cols = row.findAll("td")
            col_values = []
            for val in cols[1:]:
                col_values.append(val.text)
            key = col_values[0].split("_")
            try:
                m = Mocktest1(problem1=col_values[1],problem2=col_values[2],problem3=col_values[3],problem4=col_values[4]
                               ,total=col_values[5],student=student_dict[key[2].lower()])
                m.save()
            except KeyError as ke:
                pass
        x += 1


@cli.command()
@click.argument("student_excel_file")
@click.argument("mocktest1_html_file")
def populatedb(student_excel_file, mocktest1_html_file):
    if isExcel(student_excel_file) and isHTML(mocktest1_html_file):
        students_workbook = load_workbook(student_excel_file)
        student_worksheet = students_workbook["Current"]
        college_worksheet = students_workbook["Colleges"]
        dropped_students = students_workbook["Deletions"]
        worksheet_to_database_table(college_worksheet,student_worksheet,dropped_students, mocktest1_html_file, worksheet_startrow=2)


if __name__ == "__main__":
    cli()